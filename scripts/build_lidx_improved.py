"""개선 LIDX 모델 엑셀 산출 — 전면 수식 기반 (v3)

모든 통계량을 RawData 시트 참조 Excel 수식으로 산출.
Python 계산은 10K 시뮬레이션(Summary_Python)만 수행.
Excel 시뮬레이션(n=100)은 수식 기반으로 별도 작성.

v3 변경사항:
  - SUMPRODUCT의 --ISNUMBER → 1*ISNUMBER (XML 파싱 오류 방지)
  - COUNTIFS/AVERAGEIFS/FILTER에 전체 열 참조 사용 ($H:$H)
  - 한글 레이블 전면 적용 (DR→할인율, SG→임금상승률 등)
  - rho를 Velocity→YearlyStats로 이동
  - Simulation 시트 구조 개편 (모델별 LIDX+Y+W 통합 블록)
  - Comparison 시트 신설 (Python vs Excel 비교)

시트 구성:
  1. RawData       — 12개년 merged CSV (OK only) + helper 수식 컬럼
  2. Drift          — 연도별 drift 통계 (COUNTIFS/AVERAGEIFS/MEDIAN+FILTER)
  3. YearlyStats    — 연도별 할인율/임금상승률 + 변동 + 요약 + rho
  4. Velocity       — CV_y, CV_i (SUMPRODUCT 수식)
  5. Duration       — Duration/Convexity (AVERAGEIFS 수식)
  6. Parameters     — 모든 파라미터 셀 참조 집합
  7. Simulation     — n=100 수식 기반 시뮬레이션 (4모델)
  8. Summary_Python — n=10K Python 시뮬레이션 (기존 호환)
  9. Comparison     — Python(10K) vs Excel(100) 비교
"""

import pandas as pd
import numpy as np
from scipy import stats
from numpy.linalg import lstsq
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

print("=" * 60)
print("LIDX 개선 모델 — 전면 수식 기반 엑셀 생성 (v3)")
print("=" * 60)

wb = Workbook()
hdr_fill = PatternFill(start_color='D9E1F2', fill_type='solid')
blue_font = Font(bold=True, color='0000FF')
green_font = Font(bold=True, color='008000')
title_font = Font(bold=True, size=12)
red_font = Font(bold=True, color='FF0000')

available_years = list(range(2014, 2026))
LATEST_YEAR = 2025
N_SIM = 100
N_YEARS_SIM = 5

# ============================================================
# 1. 데이터 로드 → RawData 시트
# ============================================================
print("\n[1/9] RawData 시트 생성 중...")

all_data = []
for y in available_years:
    df = pd.read_csv(f'llm_extract_{y}_merged.csv', dtype={'corp_code': str})
    df['corp_code'] = df['corp_code'].str.zfill(8)
    ok = df[df['status'] == 'OK'].copy()
    ok['year'] = y
    all_data.append(ok)
    print(f"  {y}: {len(ok)} OK rows")

panel_full = pd.concat(all_data, ignore_index=True)
panel_full = panel_full.sort_values(['corp_code', 'year']).reset_index(drop=True)
print(f"  총 {len(panel_full)} rows")

# RawData 시트 컬럼 (A~N: 값, O~V: 수식)
RAW_COLS = [
    'year', 'corp_code', 'corp_name', 'DBO', 'ServiceCost', 'BenefitPayment',
    'InterestCost', 'DiscountRate_Mid', 'SalaryGrowth_Mid', 'Duration_Mid',
    'SensitivityDR_1pct', 'SensitivityDR_1pct_down', 'SensitivitySG_1pct',
    'SensitivitySG_1pct_down'
]
HELPER_HDRS = ['DBO_prev', 'drift(%)', '할인율_prev', '임금상승률_prev',
               'd할인율', 'd임금상승률', '유효듀레이션', '볼록성']

ws_raw = wb.active
ws_raw.title = 'RawData'

# 헤더
for ci, h in enumerate(RAW_COLS + HELPER_HDRS, 1):
    c = ws_raw.cell(row=1, column=ci, value=h)
    c.font = Font(bold=True)
    c.fill = hdr_fill

# 데이터 행 쓰기
for ri, (_, row) in enumerate(panel_full.iterrows(), 2):
    ws_raw.cell(row=ri, column=1, value=int(row['year']))          # A: year
    ws_raw.cell(row=ri, column=2, value=str(row['corp_code']))     # B: corp_code
    ws_raw.cell(row=ri, column=3, value=str(row['corp_name']))     # C: corp_name
    for ci, col in enumerate(RAW_COLS[3:], 4):                     # D~N: 수치
        v = row.get(col)
        if pd.notna(v):
            ws_raw.cell(row=ri, column=ci, value=float(v))

LAST_ROW = 1 + len(panel_full)  # 마지막 데이터 행 번호
print(f"  데이터 행: 2~{LAST_ROW} (LAST_ROW={LAST_ROW})")

# Helper 수식 (O~V) — 행 2부터
print("  Helper 수식 삽입 중...")
prev_corp = None
prev_year = None
for ri, (_, row) in enumerate(panel_full.iterrows()):
    r = ri + 2
    curr_corp = str(row['corp_code'])
    curr_year = int(row['year'])
    same_prev = (prev_corp == curr_corp and curr_year == prev_year + 1) if prev_corp else False

    # O: DBO_prev — Python에서 유효성 확인 후 직접 참조 (IF/"" 제거)
    prev_dbo_valid = same_prev and pd.notna(panel_full.iloc[ri-1].get('DBO')) and panel_full.iloc[ri-1]['DBO'] > 0
    if prev_dbo_valid:
        ws_raw.cell(row=r, column=15).value = f'=D{r-1}'

    # P: drift(%) — O/DBO/SC/BP/IC 모두 Python에서 유효 확인 후 단순 수식
    curr_dbo_valid = pd.notna(row.get('DBO')) and row['DBO'] > 0
    curr_sc_valid = pd.notna(row.get('ServiceCost'))
    curr_bp_valid = pd.notna(row.get('BenefitPayment'))
    curr_ic_valid = pd.notna(row.get('InterestCost'))
    drift_valid = prev_dbo_valid and curr_dbo_valid and curr_sc_valid and curr_bp_valid and curr_ic_valid
    if drift_valid:
        ws_raw.cell(row=r, column=16).value = f'=(E{r}-F{r}+G{r})/O{r}*100'

    # Q: 할인율_prev, R: 임금상승률_prev, S: d할인율, T: d임금상승률
    if same_prev:
        has_dr_prev = pd.notna(panel_full.iloc[ri-1].get('DiscountRate_Mid')) and panel_full.iloc[ri-1]['DiscountRate_Mid'] > 0
        has_dr_curr = pd.notna(row.get('DiscountRate_Mid')) and row['DiscountRate_Mid'] > 0
        has_sg_prev = pd.notna(panel_full.iloc[ri-1].get('SalaryGrowth_Mid')) and panel_full.iloc[ri-1]['SalaryGrowth_Mid'] > 0
        has_sg_curr = pd.notna(row.get('SalaryGrowth_Mid')) and row['SalaryGrowth_Mid'] > 0

        if has_dr_prev and has_dr_curr:
            ws_raw.cell(row=r, column=17).value = f'=H{r-1}'  # Q: 할인율_prev
            ws_raw.cell(row=r, column=19).value = f'=H{r}-H{r-1}'  # S: d할인율
        if has_sg_prev and has_sg_curr:
            ws_raw.cell(row=r, column=18).value = f'=I{r-1}'  # R: 임금상승률_prev
            ws_raw.cell(row=r, column=20).value = f'=I{r}-I{r-1}'  # T: d임금상승률

    # U: 유효듀레이션, V: 볼록성 — Python에서 범위 검증 후 단순 수식
    has_k = pd.notna(row.get('SensitivityDR_1pct'))
    has_l = pd.notna(row.get('SensitivityDR_1pct_down'))
    if has_k and has_l and curr_dbo_valid:
        # Python으로 dur_eff 범위 사전 검증
        k_val = float(row['SensitivityDR_1pct'])
        l_val = float(row['SensitivityDR_1pct_down'])
        dbo_val = float(row['DBO'])
        dur_eff_val = (l_val - k_val) / (2 * dbo_val * 0.01)
        conv_val = (l_val + k_val) / (dbo_val * 0.0001)
        if 0 < dur_eff_val < 30:
            ws_raw.cell(row=r, column=21).value = f'=(L{r}-K{r})/(2*D{r}*0.01)'
        if 0 < conv_val < 500:
            ws_raw.cell(row=r, column=22).value = f'=(L{r}+K{r})/(D{r}*0.0001)'

    prev_corp = curr_corp
    prev_year = curr_year

    if (ri + 1) % 5000 == 0:
        print(f'    {ri+1}/{len(panel_full)}')

# RawData 열 너비
for ci in range(1, 23):
    ws_raw.column_dimensions[get_column_letter(ci)].width = 14
ws_raw.column_dimensions['C'].width = 22  # corp_name

# ============================================================
# 2. Drift 시트
# ============================================================
print("\n[2/9] Drift 시트 생성 중...")
ws_drift = wb.create_sheet('Drift')

ws_drift.cell(row=1, column=1, value='[Drift = (SC-BP+IC)/DBO_{t-1}]').font = title_font
for ci, h in enumerate(['연도', '건수', 'drift 평균(%)', 'drift 중앙값(%)'], 1):
    c = ws_drift.cell(row=2, column=ci, value=h)
    c.font = Font(bold=True)
    c.fill = hdr_fill

# drift는 2015~2025 (전년 DBO 필요하므로 2014 제외)
drift_years = list(range(2015, 2026))

# 전체 열 참조 (COUNTIFS/AVERAGEIFS/FILTER용)
RD_P_COL = "RawData!$P:$P"    # drift(%)
RD_A_COL = "RawData!$A:$A"    # year
RD_H_COL = "RawData!$H:$H"    # DiscountRate_Mid
RD_I_COL = "RawData!$I:$I"    # SalaryGrowth_Mid

for di, yr in enumerate(drift_years):
    r = di + 3
    ws_drift.cell(row=r, column=1, value=yr)
    # B: 건수 — drift가 -100~100 사이인 수치만 카운트
    ws_drift.cell(row=r, column=2).value = (
        f'=COUNTIFS({RD_A_COL},A{r},{RD_P_COL},">-100",{RD_P_COL},"<100")'
    )
    # C: 평균
    ws_drift.cell(row=r, column=3).value = (
        f'=AVERAGEIFS({RD_P_COL},{RD_A_COL},A{r},{RD_P_COL},">-100",{RD_P_COL},"<100")'
    )
    # D: 중앙값 — MEDIAN(FILTER) 명시 범위 (P열에 숫자 또는 빈 셀만 존재)
    ws_drift.cell(row=r, column=4).value = (
        f'=MEDIAN(_xlfn._xlws.FILTER(RawData!$P$2:$P${LAST_ROW},'
        f'(RawData!$A$2:$A${LAST_ROW}=A{r})*(RawData!$P$2:$P${LAST_ROW}<>0)))'
    )

dr_end = 2 + len(drift_years)
SR = dr_end + 2
ws_drift.cell(row=SR, column=1, value='평균의 평균').font = blue_font
ws_drift.cell(row=SR, column=3).value = f'=AVERAGE(C3:C{dr_end})'
ws_drift.cell(row=SR, column=4).value = f'=AVERAGE(D3:D{dr_end})'

ws_drift.cell(row=SR + 2, column=1, value='[참고: 소수 변환]').font = Font(bold=True)
ws_drift.cell(row=SR + 3, column=1, value='drift_mean').font = blue_font
ws_drift.cell(row=SR + 3, column=3).value = f'=C{SR}/100'
ws_drift.cell(row=SR + 4, column=1, value='drift_median').font = blue_font
ws_drift.cell(row=SR + 4, column=3).value = f'=D{SR}/100'

# 참조 좌표 저장
DRIFT_MEAN_CELL = f"Drift!C{SR+3}"
DRIFT_MEDIAN_CELL = f"Drift!C{SR+4}"
DRIFT_MEAN_PCT_CELL = f"Drift!C{SR}"    # % 단위
DRIFT_MEDIAN_PCT_CELL = f"Drift!D{SR}"

for ci in range(1, 5):
    ws_drift.column_dimensions[get_column_letter(ci)].width = 18

print(f"  drift 연도: {drift_years[0]}~{drift_years[-1]}, 요약행: {SR}")

# ============================================================
# 3. YearlyStats 시트
# ============================================================
print("\n[3/9] YearlyStats 시트 생성 중...")
ws_ys = wb.create_sheet('YearlyStats')

ws_ys.cell(row=1, column=1, value='[연도별 할인율/임금상승률 평균]').font = title_font
hdrs = ['연도', '건수', '할인율 avg(%)', '임금상승률 avg(%)', 'Δavg_할인율(%p)', 'Δavg_임금상승률(%p)']
for ci, h in enumerate(hdrs, 1):
    c = ws_ys.cell(row=2, column=ci, value=h)
    c.font = Font(bold=True)
    c.fill = hdr_fill

for yi, yr in enumerate(available_years):
    r = yi + 3
    ws_ys.cell(row=r, column=1, value=yr)
    # B: 건수 — 할인율>0 AND 임금상승률>0
    ws_ys.cell(row=r, column=2).value = (
        f'=COUNTIFS({RD_A_COL},A{r},{RD_H_COL},">0",{RD_I_COL},">0")'
    )
    # C: 할인율 평균
    ws_ys.cell(row=r, column=3).value = (
        f'=AVERAGEIFS({RD_H_COL},{RD_A_COL},A{r},{RD_H_COL},">0",{RD_I_COL},">0")'
    )
    # D: 임금상승률 평균
    ws_ys.cell(row=r, column=4).value = (
        f'=AVERAGEIFS({RD_I_COL},{RD_A_COL},A{r},{RD_H_COL},">0",{RD_I_COL},">0")'
    )
    # E: Δavg_할인율 = C{r} - C{r-1} (첫해는 공란)
    if yi == 0:
        ws_ys.cell(row=r, column=5, value='')
        ws_ys.cell(row=r, column=6, value='')
    else:
        ws_ys.cell(row=r, column=5).value = f'=C{r}-C{r-1}'
        ws_ys.cell(row=r, column=6).value = f'=D{r}-D{r-1}'

ys_end = 2 + len(available_years)  # 마지막 데이터 행
YS_SR = ys_end + 2

ws_ys.cell(row=YS_SR, column=1, value='구분').font = Font(bold=True)
ws_ys.cell(row=YS_SR, column=3, value='할인율').font = Font(bold=True)
ws_ys.cell(row=YS_SR, column=4, value='임금상승률').font = Font(bold=True)

# YS_SR+1: 장기균형
ws_ys.cell(row=YS_SR + 1, column=1, value='장기균형(%)').font = blue_font
ws_ys.cell(row=YS_SR + 1, column=3).value = f'=AVERAGE(C3:C{ys_end})'
ws_ys.cell(row=YS_SR + 1, column=4).value = f'=AVERAGE(D3:D{ys_end})'

# YS_SR+2: 현재값
ws_ys.cell(row=YS_SR + 2, column=1, value='현재값(%)').font = blue_font
ws_ys.cell(row=YS_SR + 2, column=3).value = f'=C{ys_end}'
ws_ys.cell(row=YS_SR + 2, column=4).value = f'=D{ys_end}'

# YS_SR+3: 변동성
ws_ys.cell(row=YS_SR + 3, column=1, value='변동성(%)').font = blue_font
ws_ys.cell(row=YS_SR + 3, column=3).value = f'=STDEV(C3:C{ys_end})'
ws_ys.cell(row=YS_SR + 3, column=4).value = f'=STDEV(D3:D{ys_end})'

# YS_SR+5: 소수 변환
ws_ys.cell(row=YS_SR + 5, column=1, value='[소수 변환]').font = Font(bold=True)
decimal_items = [
    ('Y_inf', f'C{YS_SR+1}'),
    ('I_inf', f'D{YS_SR+1}'),
    ('Y_init', f'C{YS_SR+2}'),
    ('I_init', f'D{YS_SR+2}'),
    ('Vol_y', f'C{YS_SR+3}'),
    ('Vol_i', f'D{YS_SR+3}'),
]
for di, (label, ref) in enumerate(decimal_items):
    row_dec = YS_SR + 6 + di
    ws_ys.cell(row=row_dec, column=1, value=label).font = blue_font
    ws_ys.cell(row=row_dec, column=3).value = f'={ref}/100'

# rho — YearlyStats에 배치 (Velocity에서 이동)
RHO_ROW = YS_SR + 6 + len(decimal_items) + 1
ws_ys.cell(row=RHO_ROW, column=1, value='[상관계수]').font = Font(bold=True)
ws_ys.cell(row=RHO_ROW + 1, column=1, value='rho (연도평균 Δ상관)').font = blue_font
DELTA_DR_RANGE = f"E4:E{ys_end}"
DELTA_SG_RANGE = f"F4:F{ys_end}"
ws_ys.cell(row=RHO_ROW + 1, column=3).value = f'=CORREL({DELTA_DR_RANGE},{DELTA_SG_RANGE})'

RHO_CELL = f"YearlyStats!C{RHO_ROW + 1}"

# 셀 참조 좌표 저장
Y_INF_PCT_CELL = f"YearlyStats!C{YS_SR+1}"     # % 단위
I_INF_PCT_CELL = f"YearlyStats!D{YS_SR+1}"
Y_INIT_PCT_CELL = f"YearlyStats!C{YS_SR+2}"
VOL_Y_PCT_CELL = f"YearlyStats!C{YS_SR+3}"
VOL_I_PCT_CELL = f"YearlyStats!D{YS_SR+3}"
Y_INF_DEC_CELL = f"YearlyStats!C{YS_SR+6}"     # 소수 단위
Y_INIT_DEC_CELL = f"YearlyStats!C{YS_SR+8}"
VOL_Y_DEC_CELL = f"YearlyStats!C{YS_SR+10}"

for ci in range(1, 7):
    ws_ys.column_dimensions[get_column_letter(ci)].width = 18

print(f"  연도 행: 3~{ys_end}, 요약 시작: {YS_SR}, rho 행: {RHO_ROW+1}")

# ============================================================
# 4. Velocity 시트 (CV_y, CV_i — rho는 YearlyStats로 이동)
# ============================================================
print("\n[4/9] Velocity 시트 생성 중...")
ws_vel = wb.create_sheet('Velocity')

ws_vel.cell(row=1, column=1, value='[CV: 패널 MLE (X_inf 고정, 절편 없는 OLS)]').font = title_font
ws_vel.cell(row=2, column=1, value='Vasicek: ΔX(i,t) = CV × (X_inf - X(i,t-1)) + ε')

# SUMPRODUCT는 명시적 범위 유지 (전체 열은 100만 행 계산으로 극도로 느림)
RD_S = f"RawData!$S$2:$S${LAST_ROW}"    # d할인율
RD_T = f"RawData!$T$2:$T${LAST_ROW}"    # d임금상승률
RD_Q = f"RawData!$Q$2:$Q${LAST_ROW}"    # 할인율_prev
RD_R = f"RawData!$R$2:$R${LAST_ROW}"    # 임금상승률_prev

ws_vel.cell(row=4, column=1, value='파라미터').font = Font(bold=True)
ws_vel.cell(row=4, column=2, value='수식').font = Font(bold=True)
ws_vel.cell(row=4, column=3, value='값').font = Font(bold=True)

# CV_y: SUMPRODUCT(d할인율 * gap) / SUMPRODUCT(gap^2), gap = Y_inf - 할인율_prev
# 1*ISNUMBER 필터 → 0 곱으로 무효 행 제거 (--ISNUMBER 대신 XML 호환)
ws_vel.cell(row=5, column=1, value='CV_y (할인율)').font = blue_font
ws_vel.cell(row=5, column=2, value='Σ(d할인율×gap)/Σ(gap²), gap=Y_inf-할인율_prev')
ws_vel.cell(row=5, column=3).value = (
    f'=SUMPRODUCT(1*ISNUMBER({RD_S})*{RD_S}*1*ISNUMBER({RD_Q})*({Y_INF_PCT_CELL}-{RD_Q}))'
    f'/SUMPRODUCT(1*ISNUMBER({RD_Q})*(({Y_INF_PCT_CELL}-{RD_Q})^2))'
)

# CV_i
ws_vel.cell(row=6, column=1, value='CV_i (임금상승률)').font = blue_font
ws_vel.cell(row=6, column=2, value='Σ(d임금상승률×gap)/Σ(gap²), gap=I_inf-임금상승률_prev')
ws_vel.cell(row=6, column=3).value = (
    f'=SUMPRODUCT(1*ISNUMBER({RD_T})*{RD_T}*1*ISNUMBER({RD_R})*({I_INF_PCT_CELL}-{RD_R}))'
    f'/SUMPRODUCT(1*ISNUMBER({RD_R})*(({I_INF_PCT_CELL}-{RD_R})^2))'
)

# 쌍수 — SUMPRODUCT에도 명시적 범위 유지
ws_vel.cell(row=7, column=1, value='유효 쌍 수 (할인율)')
ws_vel.cell(row=7, column=3).value = f'=SUMPRODUCT((ISNUMBER({RD_S}))*1)'

ws_vel.cell(row=8, column=1, value='유효 쌍 수 (임금상승률)')
ws_vel.cell(row=8, column=3).value = f'=SUMPRODUCT((ISNUMBER({RD_T}))*1)'

CV_Y_CELL = "Velocity!C5"
CV_I_CELL = "Velocity!C6"

for ci in range(1, 4):
    ws_vel.column_dimensions[get_column_letter(ci)].width = 22
ws_vel.column_dimensions['B'].width = 55

print(f"  CV_y, CV_i 수식 삽입 완료 (rho는 YearlyStats로 이동)")

# ============================================================
# 5. Duration 시트
# ============================================================
print("\n[5/9] Duration 시트 생성 중...")
ws_dur = wb.create_sheet('Duration')

ws_dur.cell(row=1, column=1, value=f'[Duration & 볼록성 ({LATEST_YEAR})]').font = title_font
for ci, h in enumerate(['항목', '수식', '값', '건수'], 1):
    c = ws_dur.cell(row=2, column=ci, value=h)
    c.font = Font(bold=True)
    c.fill = hdr_fill

# 전체 열 참조 (COUNTIFS/AVERAGEIFS용)
RD_J_COL = "RawData!$J:$J"    # Duration_Mid
RD_U_COL = "RawData!$U:$U"    # 유효듀레이션
RD_V_COL = "RawData!$V:$V"    # 볼록성
RD_D_COL = "RawData!$D:$D"    # DBO

# --- 연도별 Duration/유효듀레이션/볼록성 추이 테이블 ---
ws_dur.cell(row=3, column=1, value='[연도별 추이]').font = Font(bold=True)
for ci, h in enumerate(['연도', 'Duration(LLM)', '건수', '유효듀레이션', '건수', '볼록성', '건수'], 1):
    c = ws_dur.cell(row=4, column=ci, value=h)
    c.font = Font(bold=True)
    c.fill = hdr_fill

for yi, yr in enumerate(available_years):
    r = yi + 5
    ws_dur.cell(row=r, column=1, value=yr)
    # B: Duration(LLM) 평균
    ws_dur.cell(row=r, column=2).value = f'=AVERAGEIFS({RD_J_COL},{RD_A_COL},A{r},{RD_J_COL},">0")'
    # C: 건수
    ws_dur.cell(row=r, column=3).value = f'=COUNTIFS({RD_A_COL},A{r},{RD_J_COL},">0")'
    # D: 유효듀레이션 평균
    ws_dur.cell(row=r, column=4).value = f'=AVERAGEIFS({RD_U_COL},{RD_A_COL},A{r},{RD_U_COL},">0")'
    # E: 건수
    ws_dur.cell(row=r, column=5).value = f'=COUNTIFS({RD_A_COL},A{r},{RD_U_COL},">0")'
    # F: 볼록성 평균
    ws_dur.cell(row=r, column=6).value = f'=AVERAGEIFS({RD_V_COL},{RD_A_COL},A{r},{RD_V_COL},">0")'
    # G: 건수
    ws_dur.cell(row=r, column=7).value = f'=COUNTIFS({RD_A_COL},A{r},{RD_V_COL},">0")'

dur_end = 4 + len(available_years)  # 마지막 데이터 행

# --- 요약 통계 ---
DUR_SR = dur_end + 2
ws_dur.cell(row=DUR_SR, column=1, value='[요약]').font = Font(bold=True)
for ci, h in enumerate(['구분', 'Duration(LLM)', '', '유효듀레이션', '', '볼록성'], 1):
    ws_dur.cell(row=DUR_SR + 1, column=ci, value=h).font = Font(bold=True)

# 전체 평균
ws_dur.cell(row=DUR_SR + 2, column=1, value='전체 평균')
ws_dur.cell(row=DUR_SR + 2, column=2).value = f'=AVERAGE(B5:B{dur_end})'
ws_dur.cell(row=DUR_SR + 2, column=4).value = f'=AVERAGE(D5:D{dur_end})'
ws_dur.cell(row=DUR_SR + 2, column=6).value = f'=AVERAGE(F5:F{dur_end})'

# 최근 3년 평균
ws_dur.cell(row=DUR_SR + 3, column=1, value='최근 3년 평균')
ws_dur.cell(row=DUR_SR + 3, column=2).value = f'=AVERAGE(B{dur_end-2}:B{dur_end})'
ws_dur.cell(row=DUR_SR + 3, column=4).value = f'=AVERAGE(D{dur_end-2}:D{dur_end})'
ws_dur.cell(row=DUR_SR + 3, column=6).value = f'=AVERAGE(F{dur_end-2}:F{dur_end})'

# 최근 5년 평균
ws_dur.cell(row=DUR_SR + 4, column=1, value='최근 5년 평균')
ws_dur.cell(row=DUR_SR + 4, column=2).value = f'=AVERAGE(B{dur_end-4}:B{dur_end})'
ws_dur.cell(row=DUR_SR + 4, column=4).value = f'=AVERAGE(D{dur_end-4}:D{dur_end})'
ws_dur.cell(row=DUR_SR + 4, column=6).value = f'=AVERAGE(F{dur_end-4}:F{dur_end})'

# 최신 연도
ws_dur.cell(row=DUR_SR + 5, column=1, value=f'최신({LATEST_YEAR})')
ws_dur.cell(row=DUR_SR + 5, column=2).value = f'=B{dur_end}'
ws_dur.cell(row=DUR_SR + 5, column=4).value = f'=D{dur_end}'
ws_dur.cell(row=DUR_SR + 5, column=6).value = f'=F{dur_end}'

# --- 선택 셀 (Parameters에서 참조) ---
SEL_ROW = DUR_SR + 7
ws_dur.cell(row=SEL_ROW, column=1, value='[사용 값 선택]').font = Font(bold=True, size=12)
ws_dur.cell(row=SEL_ROW, column=2, value='← 위 요약에서 원하는 셀 참조 입력').font = Font(color='FF0000')

ws_dur.cell(row=SEL_ROW + 1, column=1, value='Duration 선택').font = Font(bold=True, color='FF0000', size=11)
ws_dur.cell(row=SEL_ROW + 1, column=2).value = f'=AVERAGE(D{dur_end-4}:D{dur_end})'  # 유효듀레이션 최근 5년 평균
ws_dur.cell(row=SEL_ROW + 1, column=3, value='← 유효듀레이션 최근 5년 평균').font = Font(color='808080')

ws_dur.cell(row=SEL_ROW + 2, column=1, value='볼록성 선택').font = Font(bold=True, color='FF0000', size=11)
ws_dur.cell(row=SEL_ROW + 2, column=2).value = f'=AVERAGE(F{dur_end-2}:F{dur_end})'  # 볼록성 최근 3년 평균
ws_dur.cell(row=SEL_ROW + 2, column=3, value='← 볼록성 최근 3년 평균').font = Font(color='808080')

DUR_SELECTED_CELL = f"Duration!B{SEL_ROW + 1}"
CONV_SELECTED_CELL = f"Duration!B{SEL_ROW + 2}"

# 이전 호환용 참조도 유지
DUR_A_CELL = DUR_SELECTED_CELL
CONV_CELL = CONV_SELECTED_CELL
CONV_ADJ_CELL = f"Duration!B{SEL_ROW + 3}"

# 볼록성 연간 조정
ws_dur.cell(row=SEL_ROW + 3, column=1, value='볼록성 연간조정(소수)')
ws_dur.cell(row=SEL_ROW + 3, column=2).value = f'=0.5*B{SEL_ROW+2}*({VOL_Y_DEC_CELL})^2'

for ci in range(1, 8):
    ws_dur.column_dimensions[get_column_letter(ci)].width = 16
ws_dur.column_dimensions['A'].width = 20

# ============================================================
# 6. Parameters 시트
# ============================================================
print("\n[6/9] Parameters 시트 생성 중...")
ws_par = wb.create_sheet('Parameters')

ws_par.cell(row=1, column=1, value='[파라미터 요약 — 모든 값이 다른 시트 수식 참조]').font = title_font
for ci, h in enumerate(['파라미터', '의미', '값 (소수)', '값 (%)', '출처 셀'], 1):
    c = ws_par.cell(row=2, column=ci, value=h)
    c.font = Font(bold=True)
    c.fill = hdr_fill

params_list = [
    # (name, desc, dec_formula, pct_formula, source)
    ('Y_inf', '할인율 장기균형', f"={Y_INF_DEC_CELL}", f"={Y_INF_PCT_CELL}", Y_INF_PCT_CELL),
    ('Y_init', f'할인율 현재값({LATEST_YEAR})', f"={Y_INIT_DEC_CELL}", f"={Y_INIT_PCT_CELL}", Y_INIT_PCT_CELL),
    ('CV_y', '할인율 회귀속도', f"={CV_Y_CELL}", '', CV_Y_CELL),
    ('Vol_y', '할인율 변동성', f"={VOL_Y_DEC_CELL}", f"={VOL_Y_PCT_CELL}", VOL_Y_PCT_CELL),
    ('Duration', '유효듀레이션 (Duration 시트 선택)', f"={DUR_A_CELL}", '', DUR_A_CELL),
    ('볼록성', '볼록성', f"={CONV_CELL}", '', CONV_CELL),
    ('drift_mean', 'drift 보정평균', f"={DRIFT_MEAN_CELL}", f"={DRIFT_MEAN_PCT_CELL}", DRIFT_MEAN_CELL),
    ('drift_median', 'drift 보정중앙값', f"={DRIFT_MEDIAN_CELL}", f"={DRIFT_MEDIAN_PCT_CELL}", DRIFT_MEDIAN_CELL),
    ('rho', '상관계수 (참고)', f"={RHO_CELL}", '', RHO_CELL),
    ('볼록성_adj', '볼록성 연간조정', f"={CONV_ADJ_CELL}", '', CONV_ADJ_CELL),
]

for pi, (name, desc, dec_f, pct_f, src) in enumerate(params_list):
    r = pi + 3
    ws_par.cell(row=r, column=1, value=name).font = blue_font
    ws_par.cell(row=r, column=2, value=desc)
    ws_par.cell(row=r, column=3).value = dec_f
    ws_par.cell(row=r, column=4).value = pct_f if pct_f else ''
    ws_par.cell(row=r, column=5, value=src)

# Parameters 셀 참조 (Simulation에서 사용)
# C열 소수값 참조
PAR_Y_INF = "Parameters!C3"      # row 3
PAR_Y_INIT = "Parameters!C4"
PAR_CV_Y = "Parameters!C5"
PAR_VOL_Y = "Parameters!C6"
PAR_DUR = "Parameters!C7"
PAR_CONV = "Parameters!C8"
PAR_DRIFT_MEAN = "Parameters!C9"
PAR_DRIFT_MEDIAN = "Parameters!C10"
PAR_CONV_ADJ = "Parameters!C12"

for ci in range(1, 6):
    ws_par.column_dimensions[get_column_letter(ci)].width = 22

# ============================================================
# 7. Simulation 시트 (n=100, 수식 기반)
# ============================================================
print("\n[7/9] Simulation 시트 생성 중 (n=100, 4모델)...")
ws_sim = wb.create_sheet('Simulation')

# --- 난수 생성 (Python, 값으로 기록) ---
np.random.seed(42)
W_sim = np.random.randn(N_SIM, N_YEARS_SIM)

# --- Top Section ---
# Row 1: 타이틀
ws_sim.cell(row=1, column=1, value='[수식 기반 시뮬레이션 (n=100, seed=42)]').font = title_font

# Row 2~3: 파라미터 로컬 참조
ws_sim.cell(row=2, column=1, value='Y_inf').font = blue_font
ws_sim.cell(row=2, column=2).value = f'={PAR_Y_INF}'
ws_sim.cell(row=2, column=3, value='Y_init').font = blue_font
ws_sim.cell(row=2, column=4).value = f'={PAR_Y_INIT}'
ws_sim.cell(row=2, column=5, value='CV_y').font = blue_font
ws_sim.cell(row=2, column=6).value = f'={PAR_CV_Y}'
ws_sim.cell(row=2, column=7, value='Vol_y').font = blue_font
ws_sim.cell(row=2, column=8).value = f'={PAR_VOL_Y}'

ws_sim.cell(row=3, column=1, value='Duration').font = blue_font
ws_sim.cell(row=3, column=2).value = f'={PAR_DUR}'
ws_sim.cell(row=3, column=3, value='볼록성').font = blue_font
ws_sim.cell(row=3, column=4).value = f'={PAR_CONV}'
ws_sim.cell(row=3, column=5, value='drift_mean').font = blue_font
ws_sim.cell(row=3, column=6).value = f'={PAR_DRIFT_MEAN}'
ws_sim.cell(row=3, column=7, value='drift_med').font = blue_font
ws_sim.cell(row=3, column=8).value = f'={PAR_DRIFT_MEDIAN}'

# 로컬 참조 셀 (수식에서 $ 고정용)
# B2=Y_inf, D2=Y_init, F2=CV_y, H2=Vol_y
# B3=Dur, D3=Conv, F3=drift_mean, H3=drift_median

# --- 4모델 비교 테이블 (위에 배치) ---
COMP_ROW = 5
ws_sim.cell(row=COMP_ROW, column=1, value='[4모델 비교]').font = Font(bold=True, size=11)
comp_hdrs = ['모델', 'drift', '볼록성', 'Year 0', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', '1Y 증가율', '5Y 연환산']
for ci, h in enumerate(comp_hdrs, 1):
    c = ws_sim.cell(row=COMP_ROW + 1, column=ci, value=h)
    c.font = Font(bold=True)
    c.fill = hdr_fill

# We'll fill the comparison table references AFTER building the model blocks
# Placeholder rows: COMP_ROW+2 ~ COMP_ROW+5 (4 models)

# --- Per-model statistics blocks ---
STAT_SECTION_START = COMP_ROW + 7

model_defs = {
    'A': {'drift_cell': '$F$3', 'conv': False, 'label': '보정평균'},
    'B': {'drift_cell': '$F$3', 'conv': True,  'label': '보정평균+볼록성'},
    'C': {'drift_cell': '$H$3', 'conv': False, 'label': '보정중앙값'},
    'D': {'drift_cell': '$H$3', 'conv': True,  'label': '보정중앙값+볼록성'},
}

model_stats_refs = {}  # {model_key: {stat_name: row, ...}}

stat_current = STAT_SECTION_START
for mk, mdef in model_defs.items():
    ws_sim.cell(row=stat_current, column=1, value=f'[모델 {mk} 통계: {mdef["label"]}]').font = Font(bold=True)

    stat_names = ['Mean', 'Median', '5th', '25th', '75th', '95th']
    stat_rows = {}
    for si, sn in enumerate(stat_names):
        sr = stat_current + 1 + si
        stat_rows[sn] = sr
        ws_sim.cell(row=sr, column=1, value=sn).font = Font(bold=True)
        # Formulas will reference model data blocks below — fill after building data blocks

    # 5Y annualized rows
    ann_start = stat_current + 1 + len(stat_names) + 1
    for si, sn in enumerate(stat_names):
        ws_sim.cell(row=ann_start + si, column=1, value=f'{sn} 연환산').font = blue_font

    # 1Y 증가율
    yr1_row = ann_start + len(stat_names) + 1
    ws_sim.cell(row=yr1_row, column=1, value='1Y 증가율(Mean)').font = blue_font

    model_stats_refs[mk] = {
        'stat_start': stat_current,
        'stat_rows': dict(stat_rows),
        'ann_start': ann_start,
        'yr1_row': yr1_row,
    }

    stat_current = yr1_row + 2

# --- Bottom section: Model data blocks ---
# Each model: 100 rows x 17 columns
# Path# | LIDX_t0..t5 | Y_t0..t5 | W_t1..t5

DATA_SECTION_START = stat_current + 2

for mk, mdef in model_defs.items():
    BLOCK_START = DATA_SECTION_START if mk == 'A' else block_end + 3
    ws_sim.cell(row=BLOCK_START, column=1,
                value=f'[모델 {mk}: {mdef["label"]}]').font = Font(bold=True)

    # Headers row
    hdr_row = BLOCK_START + 1
    block_hdrs = ['Path#']
    for j in range(N_YEARS_SIM + 1):
        block_hdrs.append(f'LIDX_t{j}')
    for j in range(N_YEARS_SIM + 1):
        block_hdrs.append(f'Y_t{j}')
    for j in range(1, N_YEARS_SIM + 1):
        block_hdrs.append(f'W_t{j}')
    for ci, h in enumerate(block_hdrs, 1):
        c = ws_sim.cell(row=hdr_row, column=ci, value=h)
        c.font = Font(bold=True)
        c.fill = hdr_fill

    DATA_START = BLOCK_START + 2

    # Column layout:
    # A: Path#
    # B~G: LIDX_t0..t5  (cols 2..7)
    # H~M: Y_t0..t5     (cols 8..13)
    # N~R: W_t1..t5      (cols 14..18)

    for i in range(N_SIM):
        r = DATA_START + i
        ws_sim.cell(row=r, column=1, value=i + 1)

        # W columns (N~R, cols 14~18) — values from numpy
        for j in range(N_YEARS_SIM):
            ws_sim.cell(row=r, column=14 + j, value=round(float(W_sim[i, j]), 8))

        # Y columns (H~M, cols 8~13)
        # Y_t0 (col 8): Y_init
        ws_sim.cell(row=r, column=8).value = f'=$D$2'
        # Y_t1..t5 (cols 9~13): Y[t] = Y[t-1] + CV_y*(Y_inf - Y[t-1]) + Vol_y*W[t]
        for j in range(1, N_YEARS_SIM + 1):
            y_prev_col = get_column_letter(7 + j)   # previous Y column
            w_col = get_column_letter(13 + j)        # W column
            ws_sim.cell(row=r, column=8 + j).value = (
                f'={y_prev_col}{r}+$F$2*($B$2-{y_prev_col}{r})+$H$2*{w_col}{r}'
            )

        # LIDX columns (B~G, cols 2~7)
        # LIDX_t0 (col 2): 100
        ws_sim.cell(row=r, column=2, value=100.0)
        # LIDX_t1..t5 (cols 3~7)
        for j in range(1, N_YEARS_SIM + 1):
            lidx_prev_col = get_column_letter(1 + j)  # previous LIDX col
            y_cur_col = get_column_letter(8 + j)       # Y[t] col
            y_prev_col_y = get_column_letter(7 + j)    # Y[t-1] col
            dR_expr = f'({y_cur_col}{r}-{y_prev_col_y}{r})'

            if mdef['conv']:
                formula = (
                    f'={lidx_prev_col}{r}*(1+{mdef["drift_cell"]}'
                    f'-$B$3*{dR_expr}'
                    f'+0.5*$D$3*{dR_expr}^2)'
                )
            else:
                formula = (
                    f'={lidx_prev_col}{r}*(1+{mdef["drift_cell"]}'
                    f'-$B$3*{dR_expr})'
                )
            ws_sim.cell(row=r, column=2 + j).value = formula

    DATA_END = DATA_START + N_SIM - 1
    block_end = DATA_END

    # Now fill statistics formulas referencing this data block
    sref = model_stats_refs[mk]
    for sn, sr in sref['stat_rows'].items():
        for j in range(N_YEARS_SIM + 1):
            col = get_column_letter(2 + j)  # LIDX cols B~G
            rng = f'{col}{DATA_START}:{col}{DATA_END}'
            if sn == 'Mean':
                ws_sim.cell(row=sr, column=2 + j).value = f'=AVERAGE({rng})'
            elif sn == 'Median':
                ws_sim.cell(row=sr, column=2 + j).value = f'=MEDIAN({rng})'
            elif sn == '5th':
                ws_sim.cell(row=sr, column=2 + j).value = f'=PERCENTILE({rng},0.05)'
            elif sn == '25th':
                ws_sim.cell(row=sr, column=2 + j).value = f'=PERCENTILE({rng},0.25)'
            elif sn == '75th':
                ws_sim.cell(row=sr, column=2 + j).value = f'=PERCENTILE({rng},0.75)'
            elif sn == '95th':
                ws_sim.cell(row=sr, column=2 + j).value = f'=PERCENTILE({rng},0.95)'

    # 5Y annualized
    ann_start = sref['ann_start']
    stat_names_list = ['Mean', 'Median', '5th', '25th', '75th', '95th']
    for si, sn in enumerate(stat_names_list):
        sr = sref['stat_rows'][sn]
        ws_sim.cell(row=ann_start + si, column=2).value = f'=(G{sr}/100)^(1/5)-1'

    # 1Y 증가율
    yr1_row = sref['yr1_row']
    ws_sim.cell(row=yr1_row, column=2).value = f'=C{sref["stat_rows"]["Mean"]}-100'

    # Store data range for comparison sheet
    model_stats_refs[mk]['data_start'] = DATA_START
    model_stats_refs[mk]['data_end'] = DATA_END

    print(f"  모델 {mk}: data rows {DATA_START}~{DATA_END}, 통계 {sref['stat_start']}")

# --- Fill 4-model comparison table at top ---
for mi, (mk, mdef) in enumerate(model_defs.items()):
    r = COMP_ROW + 2 + mi
    sref = model_stats_refs[mk]
    ws_sim.cell(row=r, column=1, value=f'{mk}. {mdef["label"]}')
    ws_sim.cell(row=r, column=2).value = f'={mdef["drift_cell"]}'
    ws_sim.cell(row=r, column=3, value='O' if mdef['conv'] else 'X')
    # Year 0~5 (Mean)
    for j in range(N_YEARS_SIM + 1):
        mean_sr = sref['stat_rows']['Mean']
        col_letter = get_column_letter(2 + j)
        ws_sim.cell(row=r, column=4 + j).value = f'={col_letter}{mean_sr}'
    # 1Y 증가율
    ws_sim.cell(row=r, column=10).value = f'=B{sref["yr1_row"]}'
    # 5Y 연환산
    ws_sim.cell(row=r, column=11).value = f'=B{sref["ann_start"]}'

# Simulation 열 너비
for ci in range(1, 19):
    ws_sim.column_dimensions[get_column_letter(ci)].width = 14

# ============================================================
# 8. Summary_Python 시트 (n=10K Python 시뮬, 기존 호환)
# ============================================================
print("\n[8/9] Summary_Python 시트 생성 중 (n=10K Python)...")

# Python으로 파라미터 계산 (기존 코드)
dr_sg = panel_full[
    (panel_full['DiscountRate_Mid'].notna()) & (panel_full['DiscountRate_Mid'] > 0) &
    (panel_full['SalaryGrowth_Mid'].notna()) & (panel_full['SalaryGrowth_Mid'] > 0)
].copy()
dr_sg = dr_sg.rename(columns={'DiscountRate_Mid': 'DR', 'SalaryGrowth_Mid': 'SG'})

yearly_dr_sg = dr_sg.groupby('year')[['DR', 'SG']].mean()
Y_inf_pct = yearly_dr_sg['DR'].mean()
I_inf_pct = yearly_dr_sg['SG'].mean()
Vol_y_pct = yearly_dr_sg['DR'].std(ddof=1)
Vol_i_pct = yearly_dr_sg['SG'].std(ddof=1)
Y_init_pct = yearly_dr_sg['DR'].iloc[-1]
I_init_pct = yearly_dr_sg['SG'].iloc[-1]

yr_list = list(yearly_dr_sg.index)
d_dr = yearly_dr_sg['DR'].diff().dropna()
d_sg = yearly_dr_sg['SG'].diff().dropna()
valid_idx = [yr_list[i] for i in range(1, len(yr_list)) if yr_list[i] - yr_list[i - 1] == 1]
d_dr_valid = d_dr.loc[valid_idx].values
d_sg_valid = d_sg.loc[valid_idx].values
rho = np.corrcoef(d_dr_valid, d_sg_valid)[0, 1]

dr_sg_sorted = dr_sg.sort_values(['corp_name', 'year'])
dr_sg_sorted['DR_prev'] = dr_sg_sorted.groupby('corp_name')['DR'].shift(1)
dr_sg_sorted['SG_prev'] = dr_sg_sorted.groupby('corp_name')['SG'].shift(1)
dr_sg_sorted['year_prev'] = dr_sg_sorted.groupby('corp_name')['year'].shift(1)
pairs_cv = dr_sg_sorted[dr_sg_sorted['year'] - dr_sg_sorted['year_prev'] == 1].copy()
pairs_cv['dDR'] = pairs_cv['DR'] - pairs_cv['DR_prev']
pairs_cv['dSG'] = pairs_cv['SG'] - pairs_cv['SG_prev']
pairs_cv['gap_DR'] = Y_inf_pct - pairs_cv['DR_prev']
pairs_cv['gap_SG'] = I_inf_pct - pairs_cv['SG_prev']

CV_y = float(lstsq(pairs_cv['gap_DR'].values.reshape(-1, 1), pairs_cv['dDR'].values, rcond=None)[0][0])
CV_i = float(lstsq(pairs_cv['gap_SG'].values.reshape(-1, 1), pairs_cv['dSG'].values, rcond=None)[0][0])

# Drift
drift_data = []
for y in available_years:
    prev_y = y - 1
    if prev_y not in available_years:
        continue
    curr = panel_full[(panel_full['year'] == y) & (panel_full['DBO'] > 0) &
                      (panel_full['ServiceCost'].notna()) & (panel_full['BenefitPayment'].notna()) &
                      (panel_full['InterestCost'].notna())]
    prev = panel_full[(panel_full['year'] == prev_y) & (panel_full['DBO'] > 0)]
    m = curr.merge(prev[['corp_code', 'DBO']], on='corp_code', suffixes=('', '_prev'))
    m = m[m['DBO_prev'] > 0]
    d = ((m['ServiceCost'] - m['BenefitPayment'] + m['InterestCost']) / m['DBO_prev'] * 100)
    d_clean = d[d.abs() < 100]
    drift_data.append({'year': y, 'n': len(d_clean), 'mean': d_clean.mean(), 'median': d_clean.median()})

drift_mean_avg = np.mean([d['mean'] for d in drift_data])
drift_median_avg = np.mean([d['median'] for d in drift_data])

# Duration & Convexity — 유효듀레이션 최근 5년, 볼록성 최근 3년
dur_eff_yearly = []
conv_yearly = []
for y in available_years:
    df_y = panel_full[(panel_full['year'] == y) & (panel_full['DBO'] > 0) &
                      (panel_full['SensitivityDR_1pct'].notna()) &
                      (panel_full['SensitivityDR_1pct_down'].notna())].copy()
    df_y['sens_up'] = pd.to_numeric(df_y['SensitivityDR_1pct'], errors='coerce')
    df_y['sens_down'] = pd.to_numeric(df_y['SensitivityDR_1pct_down'], errors='coerce')
    df_y['dur_eff'] = (df_y['sens_down'] - df_y['sens_up']) / (2 * df_y['DBO'] * 0.01)
    df_y['conv'] = (df_y['sens_down'] + df_y['sens_up']) / (df_y['DBO'] * 0.01 ** 2)
    de = df_y[(df_y['dur_eff'] > 0) & (df_y['dur_eff'] < 30)]['dur_eff'].mean()
    cv = df_y[(df_y['conv'] > 0) & (df_y['conv'] < 500)]['conv'].mean()
    dur_eff_yearly.append(de)
    conv_yearly.append(cv)
    print(f'    {y}: 유효Dur={de:.2f}, 볼록성={cv:.1f}')

# 유효듀레이션 최근 5년 평균
Dur = np.mean(dur_eff_yearly[-5:])
# 볼록성 최근 3년 평균
Conv = np.mean(conv_yearly[-3:])
conv_adj_pct = 0.5 * Conv * (Vol_y_pct / 100) ** 2 * 100

print(f'  Y_inf={Y_inf_pct:.4f}%, Y_init={Y_init_pct:.4f}%, Vol_y={Vol_y_pct:.4f}%')
print(f'  CV_y={CV_y:.4f}, CV_i={CV_i:.4f}, rho={rho:.4f}')
print(f'  drift_mean={drift_mean_avg:.2f}%, drift_median={drift_median_avg:.2f}%')
print(f'  Duration(유효5Y)={Dur:.2f}, 볼록성(3Y)={Conv:.1f}')

# 10K 시뮬레이션
np.random.seed(42)
n_simul = 10000
n_years = 5

Y_inf = Y_inf_pct / 100
Y_init = Y_init_pct / 100
Vol_y = Vol_y_pct / 100

models = {
    'A': {'drift': drift_mean_avg / 100, 'conv': False, 'label': f'보정평균({drift_mean_avg:.1f}%)'},
    'B': {'drift': drift_mean_avg / 100, 'conv': True, 'label': f'보정평균({drift_mean_avg:.1f}%)+볼록성'},
    'C': {'drift': drift_median_avg / 100, 'conv': False, 'label': f'보정중앙값({drift_median_avg:.1f}%)'},
    'D': {'drift': drift_median_avg / 100, 'conv': True, 'label': f'보정중앙값({drift_median_avg:.1f}%)+볼록성'},
}

np.random.seed(42)
W_all = np.random.randn(n_simul, n_years)
model_results = {}

for key, cfg in models.items():
    LIDX = np.zeros((n_simul, n_years + 1))
    LIDX[:, 0] = 100.0
    Y_c = np.full(n_simul, Y_init)
    for j in range(1, n_years + 1):
        W = W_all[:, j - 1]
        Y_prev = Y_c.copy()
        Y_c = Y_prev + CV_y * (Y_inf - Y_prev) + Vol_y * W
        dR = Y_c - Y_prev
        conv_term = 0.5 * Conv * (dR ** 2) if cfg['conv'] else 0
        LIDX[:, j] = LIDX[:, j - 1] * (1 + cfg['drift'] - Dur * dR + conv_term)
    model_results[key] = LIDX

# Summary_Python 시트 작성
ws_sp = wb.create_sheet('Summary_Python')
ws_sp.cell(row=1, column=1, value='[Python 시뮬레이션 (n=10K, seed=42) — 검증용]').font = title_font

# 파라미터 요약
ws_sp.cell(row=2, column=1, value='파라미터').font = Font(bold=True)
for ci, h in enumerate(['파라미터', '의미', '산출값', '산출 방법'], 1):
    ws_sp.cell(row=3, column=ci, value=h).font = Font(bold=True)
    ws_sp.cell(row=3, column=ci).fill = hdr_fill

param_py = [
    ('Y_inf', '할인율 장기균형', round(Y_inf, 6), '연도별 평균의 평균'),
    ('Y_init', f'할인율 현재값 ({LATEST_YEAR})', round(Y_init, 6), f'{LATEST_YEAR} 연도평균'),
    ('CV_y', '할인율 회귀속도', round(CV_y, 4), '패널 MLE'),
    ('Vol_y', '할인율 변동성', round(Vol_y, 6), '연도별 평균의 std'),
    ('Duration', '유효듀레이션 (최근5Y)', round(Dur, 2), '민감도 역산, 최근 5년 평균'),
    ('볼록성', '볼록성 (최근3Y)', round(Conv, 1), '민감도 역산, 최근 3년 평균'),
    ('drift_mean', 'drift 보정평균', round(drift_mean_avg / 100, 4), '(SC-BP+IC)/DBO_{t-1} 평균'),
    ('drift_median', 'drift 보정중앙값', round(drift_median_avg / 100, 4), '(SC-BP+IC)/DBO_{t-1} 중앙값'),
    ('rho', '상관계수 (참고)', round(rho, 4), '연도평균 delta 상관'),
]

# Store Summary_Python param rows for Comparison sheet
SP_PARAM_START = 4
for pi, (name, desc, val, method) in enumerate(param_py):
    r = pi + SP_PARAM_START
    ws_sp.cell(row=r, column=1, value=name).font = Font(bold=True)
    ws_sp.cell(row=r, column=2, value=desc)
    ws_sp.cell(row=r, column=3, value=val)
    ws_sp.cell(row=r, column=4, value=method)

# 모델 비교
MT_ROW = SP_PARAM_START + len(param_py) + 2
ws_sp.cell(row=MT_ROW, column=1, value='[4모델 비교 (10K, seed=42)]').font = Font(bold=True, size=12)
headers = ['모델', 'drift', '볼록성', 'Year 0', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', '1Y 증가율', '5Y 연환산']
for i, h in enumerate(headers, 1):
    ws_sp.cell(row=MT_ROW + 1, column=i, value=h).font = Font(bold=True)
    ws_sp.cell(row=MT_ROW + 1, column=i).fill = hdr_fill

sp_model_rows = {}
for mi, (key, cfg) in enumerate(models.items()):
    r = MT_ROW + 2 + mi
    LIDX = model_results[key]
    means = [np.mean(LIDX[:, i]) for i in range(n_years + 1)]
    ann5 = (means[5] / 100) ** (1 / 5) - 1

    ws_sp.cell(row=r, column=1, value=f'{key}. {cfg["label"]}')
    ws_sp.cell(row=r, column=2, value=round(cfg['drift'] * 100, 2))
    ws_sp.cell(row=r, column=3, value='O' if cfg['conv'] else 'X')
    for yi in range(n_years + 1):
        ws_sp.cell(row=r, column=4 + yi, value=round(means[yi], 2))
    ws_sp.cell(row=r, column=10, value=round(means[1] - 100, 2))
    ws_sp.cell(row=r, column=11, value=round(ann5 * 100, 2))
    sp_model_rows[key] = r

# 분위수 테이블
sp_quantile_refs = {}  # {model_key: {stat_name: (row, col_offset)}}
for mi, (key, cfg) in enumerate(models.items()):
    LIDX = model_results[key]
    QR = MT_ROW + 8 + mi * 12
    ws_sp.cell(row=QR, column=1, value=f'[모델 {key} 분위수]').font = Font(bold=True, size=11)
    for i, h in enumerate(['통계'] + [f'Year {y}' for y in range(n_years + 1)] + ['5Y 연환산'], 1):
        ws_sp.cell(row=QR + 1, column=i, value=h).font = Font(bold=True)
        ws_sp.cell(row=QR + 1, column=i).fill = hdr_fill

    sp_quantile_refs[key] = {'qr': QR}
    for si, (sn, sf) in enumerate([
        ('Mean', np.mean), ('Stdev', np.std),
        ('5th', lambda x: np.percentile(x, 5)), ('25th', lambda x: np.percentile(x, 25)),
        ('Median', np.median), ('75th', lambda x: np.percentile(x, 75)),
        ('95th', lambda x: np.percentile(x, 95)), ('Min', np.min), ('Max', np.max),
    ]):
        ws_sp.cell(row=QR + 2 + si, column=1, value=sn)
        for yi in range(n_years + 1):
            ws_sp.cell(row=QR + 2 + si, column=2 + yi, value=round(float(sf(LIDX[:, yi])), 2))
        ann = (float(sf(LIDX[:, n_years])) / 100) ** (1 / 5) - 1
        ws_sp.cell(row=QR + 2 + si, column=2 + n_years + 1, value=round(ann * 100, 2))

# ============================================================
# 9. Comparison 시트
# ============================================================
print("\n[9/9] Comparison 시트 생성 중...")
ws_cmp = wb.create_sheet('Comparison')

ws_cmp.cell(row=1, column=1, value='[Python(n=10K) vs Excel(n=100) 비교]').font = title_font
for ci, h in enumerate(['항목', 'Python(n=10K)', 'Excel(n=100)', '차이'], 1):
    c = ws_cmp.cell(row=2, column=ci, value=h)
    c.font = Font(bold=True)
    c.fill = hdr_fill

cmp_row = 3

# 모델별 1Y 증가율 비교
ws_cmp.cell(row=cmp_row, column=1, value='[모델별 성과]').font = Font(bold=True)
cmp_row += 1

for mk in ['A', 'B', 'C', 'D']:
    sref = model_stats_refs[mk]
    sp_r = sp_model_rows[mk]

    # 1Y 증가율
    ws_cmp.cell(row=cmp_row, column=1, value=f'모델 {mk} 1Y증가율')
    ws_cmp.cell(row=cmp_row, column=2).value = f'=Summary_Python!J{sp_r}'
    ws_cmp.cell(row=cmp_row, column=3).value = f'=Simulation!B{sref["yr1_row"]}'
    ws_cmp.cell(row=cmp_row, column=4).value = f'=C{cmp_row}-B{cmp_row}'
    cmp_row += 1

    # 5Y 연환산
    ws_cmp.cell(row=cmp_row, column=1, value=f'모델 {mk} 5Y연환산')
    ws_cmp.cell(row=cmp_row, column=2).value = f'=Summary_Python!K{sp_r}'
    ws_cmp.cell(row=cmp_row, column=3).value = f'=Simulation!B{sref["ann_start"]}*100'
    ws_cmp.cell(row=cmp_row, column=4).value = f'=C{cmp_row}-B{cmp_row}'
    cmp_row += 1

cmp_row += 1

# 파라미터 비교
ws_cmp.cell(row=cmp_row, column=1, value='[파라미터]').font = Font(bold=True)
cmp_row += 1

param_compare = [
    ('Y_inf', 0),
    ('Y_init', 1),
    ('CV_y', 2),
    ('Vol_y', 3),
    ('Duration', 4),
    ('볼록성', 5),
    ('drift_mean', 6),
    ('drift_median', 7),
    ('rho', 8),
]

for pname, pi in param_compare:
    sp_param_row = SP_PARAM_START + pi
    par_row = 3 + pi  # Parameters sheet row
    ws_cmp.cell(row=cmp_row, column=1, value=pname)
    ws_cmp.cell(row=cmp_row, column=2).value = f'=Summary_Python!C{sp_param_row}'
    ws_cmp.cell(row=cmp_row, column=3).value = f'=Parameters!C{par_row}'
    ws_cmp.cell(row=cmp_row, column=4).value = f'=C{cmp_row}-B{cmp_row}'
    cmp_row += 1

for ci in range(1, 5):
    ws_cmp.column_dimensions[get_column_letter(ci)].width = 20

# ============================================================
# 서식 정리 & 저장
# ============================================================
print("\n열 너비 조정 & 저장...")

for ws in [ws_drift, ws_ys, ws_vel, ws_dur, ws_par, ws_sp, ws_cmp]:
    if ws.title not in ('RawData', 'Simulation', 'Velocity'):
        for col in range(1, 30):
            ws.column_dimensions[get_column_letter(col)].width = 18

out_file = 'lidx_improved.xlsx'
wb.save(out_file)

# ── XML 후처리: 수식 셀의 빈 <v></v> 태그 제거 (Excel 복구 팝업 방지) ──
print("XML 후처리 (빈 <v></v> 제거)...")
import zipfile, shutil, tempfile, re as re_mod

tmp_file = out_file + '.tmp'
with zipfile.ZipFile(out_file, 'r') as zin:
    with zipfile.ZipFile(tmp_file, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                content = data.decode('utf-8')
                # 수식 뒤의 빈 <v></v> 제거: <f>...</f><v></v> → <f>...</f>
                content = re_mod.sub(r'(</f>)<v></v>', r'\1', content)
                data = content.encode('utf-8')
                item.file_size = len(data)
                item.compress_size = len(data)  # will be recalculated
            zout.writestr(item, data)

shutil.move(tmp_file, out_file)
print(f"  완료 — 빈 <v></v> 제거됨")

print(f"\n{'=' * 60}")
print(f"{out_file} 저장 완료")
print(f"{'=' * 60}")
print(f"\n시트 구조:")
print(f"  1. RawData        — {LAST_ROW - 1}행, 22컬럼 (A~N 값 + O~V 수식)")
print(f"  2. Drift           — drift 통계 (COUNTIFS/AVERAGEIFS/MEDIAN+FILTER)")
print(f"  3. YearlyStats     — 할인율/임금상승률 연도별 (수식) + rho")
print(f"  4. Velocity        — CV_y/CV_i (SUMPRODUCT 수식)")
print(f"  5. Duration        — Duration/볼록성 (AVERAGEIFS 수식)")
print(f"  6. Parameters      — 모든 파라미터 셀 참조")
print(f"  7. Simulation      — n=100 수식 시뮬 (4모델, 통합 블록)")
print(f"  8. Summary_Python  — n=10K Python 시뮬 (검증용)")
print(f"  9. Comparison      — Python vs Excel 비교")
